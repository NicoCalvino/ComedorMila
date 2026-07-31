"""Utilidad para exportar reportes a Excel (.xlsx) con openpyxl."""

import re
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_AZUL = "004282"


def xlsx_response(filename, headers, rows, titulo=None):
    """Devuelve un HttpResponse con un .xlsx armado a partir de headers + rows.

    - headers: lista de títulos de columna.
    - rows: lista de filas (cada fila una lista de valores).
    - titulo: opcional, se escribe arriba en negrita.
    """
    wb = Workbook()
    ws = wb.active
    # El nombre de la hoja no admite / \ ? * [ ] : ni más de 31 chars.
    ws.title = re.sub(r'[\\/?*\[\]:]', '-', (titulo or "Reporte"))[:31]

    fila = 1
    if titulo:
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14, color=_AZUL)
        fila = 3

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_AZUL)
    center = Alignment(horizontal="center", vertical="center")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=fila, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for i, row in enumerate(rows, start=fila + 1):
        for j, val in enumerate(row, start=1):
            if isinstance(val, Decimal):
                val = float(val)
            ws.cell(row=i, column=j, value=val)

    # Ancho de columnas según el contenido.
    for j, h in enumerate(headers, start=1):
        largo = len(str(h))
        for row in rows:
            largo = max(largo, len(str(row[j - 1])))
        ws.column_dimensions[get_column_letter(j)].width = min(max(largo + 2, 10), 48)

    ws.freeze_panes = ws.cell(row=fila + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)  # openpyxl necesita un stream con seek; HttpResponse no lo tiene
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
